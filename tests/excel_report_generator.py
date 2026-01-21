"""
Excel Report Generator
Collects all report data and generates Excel files with charts and comparisons.
"""

import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import logging

try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Missing required packages. Please install: pip install pandas openpyxl")
    raise

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates and updates Excel reports with charts from crypto analysis data."""
    
    def __init__(self, reports_dir: str = None, output_dir: str = None):
        """
        Initialize Excel report generator.
        
        Args:
            reports_dir: Directory containing report folders
            output_dir: Directory to save Excel files
        """
        if reports_dir is None:
            reports_dir = Path(__file__).parent.parent / "data" / "reports"
        if output_dir is None:
            output_dir = Path(__file__).parent
            
        self.reports_dir = Path(reports_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def parse_report(self, report_path: Path) -> Optional[Dict]:
        """Parse a single report JSON file."""
        try:
            with open(report_path, 'r') as f:
                data = json.load(f)
            return data
        except Exception as e:
            logger.error(f"Error parsing {report_path}: {e}")
            return None
    
    def collect_symbol_reports(self, symbol: str) -> List[Dict]:
        """
        Collect all reports for a specific symbol.
        
        Args:
            symbol: Cryptocurrency symbol (e.g., BTCUSDT, SOLUSDT)
            
        Returns:
            List of report data dictionaries
        """
        reports = []
        
        # Scan all report directories
        for report_dir in sorted(self.reports_dir.iterdir()):
            if not report_dir.is_dir():
                continue
                
            # Check if directory name matches symbol
            if not report_dir.name.startswith(symbol):
                continue
                
            report_file = report_dir / "report.json"
            if not report_file.exists():
                continue
                
            report_data = self.parse_report(report_file)
            if report_data:
                reports.append(report_data)
        
        # Sort by timestamp
        reports.sort(key=lambda x: x['metadata']['analysis_timestamp'])
        
        logger.info(f"Collected {len(reports)} reports for {symbol}")
        return reports
    
    def extract_data_for_excel(self, reports: List[Dict]) -> pd.DataFrame:
        """
        Extract relevant data from reports into a DataFrame.
        
        Args:
            reports: List of report dictionaries
            
        Returns:
            DataFrame with all extracted data
        """
        data_rows = []
        
        for report in reports:
            try:
                metadata = report.get('metadata', {})
                market = report.get('market_data', {})
                fundamental = report.get('fundamental_analysis', {})
                technical = report.get('technical_analysis', {})
                sentiment = report.get('sentiment_analysis', {})
                predictions = report.get('predictions', {})
                recommendations = report.get('recommendations', {})
                
                # Parse and format timestamp
                timestamp_str = metadata.get('analysis_timestamp', '')
                dt = pd.to_datetime(timestamp_str)
                formatted_timestamp = dt.strftime('%H:%M:%S %d/%m/%Y') if pd.notna(dt) else ''
                
                row = {
                    # Timestamp
                    'timestamp': formatted_timestamp,
                    'datetime': dt,
                    
                    # Market Data
                    'current_price': market.get('current_price', 0),
                    '24h_high': market.get('high_24h', 0),
                    '24h_low': market.get('low_24h', 0),
                    '24h_change': market.get('price_change_24h', 0),
                    '24h_change_percent': market.get('price_change_percent_24h', 0),
                    '24h_volume': market.get('volume_24h', 0),
                    '24h_quote_volume': market.get('quote_volume_24h', 0),
                    '24h_trades': market.get('trades_24h', 0),
                    
                    # Fundamental Analysis
                    'fundamental_score': fundamental.get('overall_score', 0),
                    'tokenomics_score': fundamental.get('tokenomics_score', 0),
                    'team_score': fundamental.get('team_score', 0),
                    'technology_score': fundamental.get('technology_score', 0),
                    'adoption_score': fundamental.get('adoption_score', 0),
                    
                    # Technical Analysis
                    'trend_direction': technical.get('trend_direction', ''),
                    'trend_strength': technical.get('trend_strength', 0),
                    'momentum_score': technical.get('momentum_score', 0),
                    'volatility': technical.get('volatility_assessment', ''),
                    'rsi': technical.get('indicators', {}).get('rsi', {}).get('value', 0),
                    'rsi_signal': technical.get('indicators', {}).get('rsi', {}).get('signal', ''),
                    'macd': technical.get('indicators', {}).get('macd', {}).get('value', 0),
                    'macd_signal': technical.get('indicators', {}).get('macd', {}).get('signal', ''),
                    
                    # Sentiment
                    'overall_sentiment': sentiment.get('overall_sentiment', 0),
                    'news_sentiment': sentiment.get('news_sentiment', 0),
                    'social_sentiment': sentiment.get('social_sentiment', 0),
                    'fear_greed_score': sentiment.get('market_fear_greed', {}).get('score', 0),
                    'fear_greed_category': sentiment.get('market_fear_greed', {}).get('category', ''),
                    
                    # Predictions
                    'probability_up': predictions.get('probability_up', 0),
                    'probability_down': predictions.get('probability_down', 0),
                    'expected_volatility': predictions.get('expected_volatility', 0),
                    'forecasted_price': predictions.get('price_forecast', {}).get('forecasted_price', 0),
                    'expected_change': predictions.get('price_forecast', {}).get('expected_change', 0),
                    'forecast_trend': predictions.get('price_forecast', {}).get('trend_direction', ''),
                    
                    # Support/Resistance
                    'nearest_support': predictions.get('next_24h_support_levels', [{}])[0].get('level', 0) if predictions.get('next_24h_support_levels') else 0,
                    'nearest_resistance': predictions.get('next_24h_resistance_levels', [{}])[0].get('level', 0) if predictions.get('next_24h_resistance_levels') else 0,
                    
                    # Recommendations
                    'stop_loss': recommendations.get('stop_loss_suggestion', 0),
                    'recommendation_confidence': recommendations.get('confidence', 0),
                    'trading_suggestion': recommendations.get('mock_trading_suggestion', ''),
                    
                    # Overall Confidence
                    'confidence_score': metadata.get('confidence_score', 0),
                }
                
                data_rows.append(row)
            except Exception as e:
                logger.error(f"Error extracting data from report: {e}")
                continue
        
        df = pd.DataFrame(data_rows)
        
        # Sort by datetime
        if not df.empty and 'datetime' in df.columns:
            df = df.sort_values('datetime')
        
        return df
    
    def create_excel_with_charts(self, symbol: str, df: pd.DataFrame) -> str:
        """
        Create or update Excel file with data and charts.
        
        Args:
            symbol: Cryptocurrency symbol
            df: DataFrame with report data
            
        Returns:
            Path to created/updated Excel file
        """
        if df.empty:
            logger.warning(f"No data to export for {symbol}")
            return None
            
        excel_path = self.output_dir / f"{symbol}_analysis.xlsx"
        
        # Create new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Create Data Sheet
        ws_data = wb.create_sheet("Data")
        
        # Write headers with styling
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # 2. Create Price Analysis Sheet with Charts
        self._create_price_analysis_sheet(wb, df)
        
        # 3. Create Technical Indicators Sheet
        self._create_technical_sheet(wb, df)
        
        # 4. Create Fundamental Analysis Sheet
        self._create_fundamental_sheet(wb, df)
        
        # 5. Create Sentiment Analysis Sheet
        self._create_sentiment_sheet(wb, df)
        
        # 6. Create Predictions Sheet
        self._create_predictions_sheet(wb, df)
        
        # 7. Create Summary Dashboard
        self._create_summary_dashboard(wb, df, symbol)
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Excel report saved to: {excel_path}")
        
        return str(excel_path)
    
    def _create_price_analysis_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create price analysis sheet with charts."""
        ws = wb.create_sheet("Price Analysis")
        
        # Add title
        ws['A1'] = "Price Analysis & 24-Hour Trading Range"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Write key metrics
        row = 3
        metrics = ['timestamp', 'datetime', 'current_price', '24h_high', '24h_low', '24h_change', '24h_change_percent', '24h_volume']
        headers = ['Timestamp', 'DateTime', 'Current Price', '24h High', '24h Low', '24h Change', '24h Change %', '24h Volume']
        
        for col, (metric, header) in enumerate(zip(metrics, headers), 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            for idx, value in enumerate(df[metric], 1):
                ws.cell(row=row + idx, column=col, value=value)
        
        # Create Price Chart
        chart = LineChart()
        chart.title = "Price Movement"
        chart.style = 10
        chart.height = 10
        chart.width = 20
        chart.y_axis.title = "Price (USD)"
        chart.x_axis.title = "Time"
        
        data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=5)
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "I3")
        
        # Create Volume Chart
        volume_chart = BarChart()
        volume_chart.title = "24-Hour Trading Volume"
        volume_chart.style = 10
        volume_chart.height = 10
        volume_chart.width = 20
        volume_chart.y_axis.title = "Volume"
        
        vol_data = Reference(ws, min_col=8, min_row=row, max_row=row + len(df))
        volume_chart.add_data(vol_data, titles_from_data=True)
        volume_chart.set_categories(cats)
        
        ws.add_chart(volume_chart, "I23")
    
    def _create_technical_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create technical analysis sheet with charts."""
        ws = wb.create_sheet("Technical Analysis")
        
        ws['A1'] = "Technical Indicators"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        metrics = ['timestamp', 'datetime', 'trend_strength', 'momentum_score', 'rsi', 'macd']
        headers = ['Timestamp', 'DateTime', 'Trend Strength', 'Momentum', 'RSI', 'MACD']
        
        for col, (metric, header) in enumerate(zip(metrics, headers), 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            for idx, value in enumerate(df[metric], 1):
                ws.cell(row=row + idx, column=col, value=value)
        
        # RSI Chart
        chart = LineChart()
        chart.title = "RSI Indicator"
        chart.y_axis.title = "RSI Value"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=5, min_row=row, max_row=row + len(df))
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
        
        # Trend & Momentum Chart
        trend_chart = LineChart()
        trend_chart.title = "Trend Strength & Momentum"
        trend_chart.y_axis.title = "Score"
        trend_chart.height = 10
        trend_chart.width = 20
        
        trend_data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=4)
        trend_chart.add_data(trend_data, titles_from_data=True)
        trend_chart.set_categories(cats)
        
        ws.add_chart(trend_chart, "G23")
    
    def _create_fundamental_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create fundamental analysis sheet."""
        ws = wb.create_sheet("Fundamental Analysis")
        
        ws['A1'] = "Fundamental Scores"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        metrics = ['timestamp', 'datetime', 'fundamental_score', 'tokenomics_score', 'team_score', 'technology_score', 'adoption_score']
        headers = ['Timestamp', 'DateTime', 'Overall Score', 'Tokenomics', 'Team', 'Technology', 'Adoption']
        
        for col, (metric, header) in enumerate(zip(metrics, headers), 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            for idx, value in enumerate(df[metric], 1):
                ws.cell(row=row + idx, column=col, value=value)
        
        # Create chart
        chart = LineChart()
        chart.title = "Fundamental Analysis Scores"
        chart.y_axis.title = "Score"
        chart.height = 12
        chart.width = 24
        
        data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=7)
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "H3")
    
    def _create_sentiment_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create sentiment analysis sheet."""
        ws = wb.create_sheet("Sentiment Analysis")
        
        ws['A1'] = "Sentiment Indicators"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        metrics = ['timestamp', 'datetime', 'overall_sentiment', 'news_sentiment', 'social_sentiment', 'fear_greed_score']
        headers = ['Timestamp', 'DateTime', 'Overall Sentiment', 'News Sentiment', 'Social Sentiment', 'Fear & Greed']
        
        for col, (metric, header) in enumerate(zip(metrics, headers), 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            for idx, value in enumerate(df[metric], 1):
                ws.cell(row=row + idx, column=col, value=value)
        
        # Sentiment Chart
        chart = LineChart()
        chart.title = "Sentiment Analysis"
        chart.y_axis.title = "Sentiment Score"
        chart.height = 12
        chart.width = 24
        
        data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=5)
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
        
        # Fear & Greed Chart
        fg_chart = LineChart()
        fg_chart.title = "Fear & Greed Index"
        fg_chart.y_axis.title = "Index Value"
        fg_chart.height = 10
        fg_chart.width = 20
        
        fg_data = Reference(ws, min_col=6, min_row=row, max_row=row + len(df))
        fg_chart.add_data(fg_data, titles_from_data=True)
        fg_chart.set_categories(cats)
        
        ws.add_chart(fg_chart, "G23")
    
    def _create_predictions_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create predictions sheet."""
        ws = wb.create_sheet("Predictions")
        
        ws['A1'] = "Price Predictions & Probabilities"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        metrics = ['timestamp', 'datetime', 'current_price', 'forecasted_price', 'expected_change', 
                   'probability_up', 'probability_down', 'nearest_support', 'nearest_resistance']
        headers = ['Timestamp', 'DateTime', 'Current Price', 'Forecasted Price', 'Expected Change',
                   'Probability Up', 'Probability Down', 'Nearest Support', 'Nearest Resistance']
        
        for col, (metric, header) in enumerate(zip(metrics, headers), 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            for idx, value in enumerate(df[metric], 1):
                ws.cell(row=row + idx, column=col, value=value)
        
        # Price Forecast Chart
        chart = LineChart()
        chart.title = "Current vs Forecasted Price"
        chart.y_axis.title = "Price (USD)"
        chart.height = 12
        chart.width = 24
        
        data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=4)
        cats = Reference(ws, min_col=2, min_row=row + 1, max_row=row + len(df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "J3")
        
        # Support/Resistance Chart
        sr_chart = LineChart()
        sr_chart.title = "Support & Resistance Levels"
        sr_chart.y_axis.title = "Price (USD)"
        sr_chart.height = 10
        sr_chart.width = 20
        
        sr_data = Reference(ws, min_col=3, min_row=row, max_row=row + len(df), max_col=3)
        sr_levels = Reference(ws, min_col=8, min_row=row, max_row=row + len(df), max_col=9)
        sr_chart.add_data(sr_data, titles_from_data=True)
        sr_chart.add_data(sr_levels, titles_from_data=True)
        sr_chart.set_categories(cats)
        
        ws.add_chart(sr_chart, "J23")
    
    def _create_summary_dashboard(self, wb: Workbook, df: pd.DataFrame, symbol: str):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Dashboard", 0)  # Insert as first sheet
        
        # Title
        ws['A1'] = f"{symbol} Analysis Dashboard"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Latest metrics
        if not df.empty:
            latest = df.iloc[-1]
            
            row = 3
            ws[f'A{row}'] = "Latest Analysis"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 1
            metrics = [
                ("Timestamp", latest['timestamp']),
                ("Current Price", f"${latest['current_price']:.2f}"),
                ("24h Change", f"{latest['24h_change_percent']:.2f}%"),
                ("Forecasted Price", f"${latest['forecasted_price']:.2f}"),
                ("Trend", latest['trend_direction']),
                ("Overall Sentiment", f"{latest['overall_sentiment']:.3f}"),
                ("Confidence Score", f"{latest['confidence_score']:.2f}"),
                ("Trading Suggestion", latest['trading_suggestion']),
            ]
            
            for label, value in metrics:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            # Summary statistics
            row += 2
            ws[f'A{row}'] = "Summary Statistics"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 1
            stats = [
                ("Total Reports", len(df)),
                ("Average Price", f"${df['current_price'].mean():.2f}"),
                ("Price Range", f"${df['current_price'].min():.2f} - ${df['current_price'].max():.2f}"),
                ("Average Volatility", f"{df['expected_volatility'].mean():.2f}%"),
                ("Average RSI", f"{df['rsi'].mean():.2f}"),
                ("Average Sentiment", f"{df['overall_sentiment'].mean():.3f}"),
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def update_excel_with_new_report(self, symbol: str, report_data: Dict) -> str:
        """
        Update existing Excel file with new report data or create new if doesn't exist.
        
        Args:
            symbol: Cryptocurrency symbol
            report_data: New report data dictionary
            
        Returns:
            Path to updated Excel file
        """
        # Collect all reports including the new one
        reports = self.collect_symbol_reports(symbol)
        
        # Extract data and create/update Excel
        df = self.extract_data_for_excel(reports)
        excel_path = self.create_excel_with_charts(symbol, df)
        
        return excel_path
    
    def generate_all_reports(self) -> List[str]:
        """
        Generate Excel reports for all symbols found in reports directory.
        
        Returns:
            List of paths to created Excel files
        """
        # Find all unique symbols
        symbols = set()
        for report_dir in self.reports_dir.iterdir():
            if not report_dir.is_dir():
                continue
            # Extract symbol from directory name (format: SYMBOL_TIMESTAMP)
            parts = report_dir.name.split('_')
            if len(parts) >= 2:
                symbol = parts[0]
                symbols.add(symbol)
        
        excel_files = []
        for symbol in sorted(symbols):
            logger.info(f"Generating report for {symbol}...")
            reports = self.collect_symbol_reports(symbol)
            if reports:
                df = self.extract_data_for_excel(reports)
                excel_path = self.create_excel_with_charts(symbol, df)
                if excel_path:
                    excel_files.append(excel_path)
        
        return excel_files


def main():
    """Main function for standalone execution."""
    import sys
    
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    generator = ExcelReportGenerator()
    
    if len(sys.argv) > 1:
        # Generate for specific symbol
        symbol = sys.argv[1]
        print(f"\n📊 Generating Excel report for {symbol}...")
        reports = generator.collect_symbol_reports(symbol)
        if reports:
            df = generator.extract_data_for_excel(reports)
            excel_path = generator.create_excel_with_charts(symbol, df)
            print(f"✅ Report saved to: {excel_path}")
        else:
            print(f"❌ No reports found for {symbol}")
    else:
        # Generate for all symbols
        print("\n📊 Generating Excel reports for all symbols...")
        excel_files = generator.generate_all_reports()
        print(f"\n✅ Generated {len(excel_files)} Excel reports:")
        for path in excel_files:
            print(f"   - {path}")


if __name__ == "__main__":
    main()
